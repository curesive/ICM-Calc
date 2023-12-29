#! python3
#Monte Carlo Sim operating on a single tournament
#Using Sheet1 from TournamentData.xlsx
#Output goes to updatedTournamentData.xlsx

from random import *
import openpyxl, sys
from openpyxl.styles import Font

def trial(weights, payouts):           
        #scores is a list, the items inside the list are tuples. Each tuple contains the random score
        #and it's corresponding stack number. [(0.15533, 8), (0.18223, 3) ... ]
        #low weight = best chance of winning
        scores = sorted((random() ** weight, i) for i, weight in enumerate(weights))

        results = [0] * len(payouts)
        for payout, score in zip(payouts, scores): results[score[1]] = payout
        return results

def sicm(wbName):
        wb = openpyxl.load_workbook(wbName)
        sheet = wb['Sheet1']

        stacks = [0]*2000
        for i in range(0,1999):
                stacks[i] = sheet['B' + str(i+4)].value
        #get rid of the blank values in stacks
        stacks = list(filter(None, stacks))
     

        avg = sum(stacks) / float(len(stacks)) #avg is average stack size
        weights = [avg / s for s in stacks] #list with items equal to the average stack divided by each individual stack
                      
                  
        return [sum(player) / float(globalTrials) for player in zip(*( 
                trial(weights, globalPayouts) for i in range(globalTrials) 
        ))]

def printDist(wbName, icmValues):
        wb = openpyxl.load_workbook(wbName)
        sheet = wb['Sheet1']
                     

        for i in range(0, len(icmValues)):
                sheet.cell(row = i + 4, column = 3).value = icmValues[i]

        wb.save('updatedTournamentData.xlsx')
        return

if __name__ == '__main__':
        wb = openpyxl.load_workbook('TournamentData.xlsx')
        sheet = wb['Sheet1']        
        globalTrials = sheet['C1'].value
                 
        
        payoutList = []
        finishDistribution = []
        globalFinishPercentage = []
        

        globalPayouts = [0]*2000
        for i in range(0,1999):
                globalPayouts[i] = sheet['D' + str(i+4)].value
        globalPayouts = list(filter(None, globalPayouts))
        globalPayouts.sort() #arranges values in payouts from lowest number to highest

        payoutList.append({})
        finishDistribution.append({})        
        
        icmValues = sicm('TournamentData.xlsx')
        printDist('TournamentData.xlsx', icmValues)
        sys.exit()

       
        

        

